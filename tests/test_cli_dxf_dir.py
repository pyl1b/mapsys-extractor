"""CLI tests for the directory-wide DXF conversion command."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from click.testing import CliRunner

from mapsys.cli import cli


def _make_minimal_template(tmp_path: Path) -> Path:
    """Create a minimal DXF template with POINT block."""
    import ezdxf

    doc = ezdxf.new(setup=True)
    blk = doc.blocks.new("POINT")
    blk.add_attdef("NAME", insert=(0, 0), height=0.2)
    blk.add_attdef("SOURCE", insert=(0, 0), height=0.2)
    blk.add_attdef("Z", insert=(0, 0), height=0.2)
    path = tmp_path / "template.dxf"
    doc.saveas(path.as_posix())
    return path


class _DummyContent:
    """Minimal Content-like object for testing."""

    def __init__(self, main_file: Path) -> None:
        self.main_file = main_file
        self.points = [object()] * 3
        self.texts = [object()] * 2
        self.t_meta = [object()]
        self.p_meta = [object()]
        self.v_offsets = [0, 1, 2]
        self.p_layers = [object()]
        self.pr5 = _DummyPr5()

    def text_by_offset(self, offset: int) -> str | None:
        return "TXT" if offset == 0 else None

    def get_poly_layer(self, p_meta: Any) -> int:
        return 0


class _DummyPr5:
    """Minimal PR5-like object for testing."""

    def __init__(self) -> None:
        self.layers = [object(), object()]
        self.after: list[object] = []
        self.font_names = ["Arial"]


def test_to_dxf_dir_with_report_creates_excel(
    tmp_path: Path, monkeypatch: Any
) -> None:
    """`to-dxf-dir --report` creates an Excel report with conversion data."""
    template = _make_minimal_template(tmp_path)
    (tmp_path / "proj.pr5").write_bytes(b"x")

    def _fake_create(main_file: Path) -> Any:
        return _DummyContent(main_file)

    def _fake_convert(
        content: Any,
        dxf_template: Path,
        *,
        dxf_path: Path | None = None,
        **kwargs: Any,
    ) -> Any:
        if dxf_path is not None:
            dxf_path.write_bytes(b"0" * 200)
        return None

    import mapsys.dxf.to_dxf as mto_dxf
    import mapsys.parser.content as mcontent

    monkeypatch.setattr(mcontent.Content, "create", _fake_create)
    monkeypatch.setattr(mto_dxf.Builder, "convert", _fake_convert)

    report_path = tmp_path / "conversion_report.xlsx"
    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "to-dxf-dir",
            str(tmp_path),
            "--dxf-template",
            str(template),
            "--report",
            str(report_path),
        ],
    )

    assert result.exit_code == 0
    assert report_path.exists()
    assert "Report written" in result.output
    assert (tmp_path / "proj.dxf").exists()

    # Verify report structure.
    from openpyxl import load_workbook

    wb = load_workbook(report_path)
    assert "Conversion report" in wb.sheetnames
    ws = wb["Conversion report"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "Mapsys path" in headers
    assert "points" in headers
    assert ws.max_row >= 2
    assert ws.cell(row=2, column=3).value == "proj"
    assert ws.cell(row=2, column=4).value == 3
    assert ws.cell(row=2, column=14).value == 200


def test_to_dxf_dir_report_empty_when_no_conversions(
    tmp_path: Path, monkeypatch: Any
) -> None:
    """`to-dxf-dir --report` creates report with headers only when no .pr5."""
    template = _make_minimal_template(tmp_path)
    report_path = tmp_path / "report.xlsx"
    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "to-dxf-dir",
            str(tmp_path),
            "--dxf-template",
            str(template),
            "--report",
            str(report_path),
        ],
    )
    assert result.exit_code == 0
    assert report_path.exists()
    from openpyxl import load_workbook

    wb = load_workbook(report_path)
    ws = wb["Conversion report"]
    assert ws.max_row == 1
