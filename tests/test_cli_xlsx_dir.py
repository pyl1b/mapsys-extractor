"""CLI tests for the directory-wide XLSX export command."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from click.testing import CliRunner

from mapsys.cli import cli


class _DummyContent:
    def __init__(self) -> None:
        self.points: list[object] = []
        self.texts: list[object] = []
        self.t_meta: list[object] = []
        self.p_meta: list[object] = []
        self.v_offsets: list[int] = []
        self.p_layers: list[object] = []
        self.pr5: object | None = None


def test_cli_to_xlsx_dir_exports_all(tmp_path: Path, monkeypatch: Any) -> None:
    """`to-xlsx-dir` writes XLSX files for each .pr5 and prints a summary."""

    # Two projects: one in root, one in a subdir
    (tmp_path / "a.pr5").write_bytes(b"a")
    sub = tmp_path / "sub"
    sub.mkdir()
    (sub / "b.pr5").write_bytes(b"b")

    def _fake_create(_: Path) -> Any:
        return _DummyContent()

    import mapsys.parser.content as mcontent

    monkeypatch.setattr(mcontent.Content, "create", _fake_create)

    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "to-xlsx-dir",
            str(tmp_path),
            "--max-depth",
            "1",
        ],
    )

    assert result.exit_code == 0
    # Summary contains numbers of processed directories and exported files
    assert "exported" in result.output
    # Both XLSX should be created next to each .pr5
    assert (tmp_path / "a.xlsx").exists()
    assert (sub / "b.xlsx").exists()


def test_cli_to_xlsx_dir_with_report_creates_excel(
    tmp_path: Path, monkeypatch: Any
) -> None:
    """`to-xlsx-dir --report` creates an Excel report with export data."""
    (tmp_path / "proj.pr5").write_bytes(b"x")

    class _Pr5Like:
        layers = [object()]
        after: list[object] = []
        font_names = ["Arial", "Calibri"]

    class _ContentWithPr5(_DummyContent):
        def __init__(self, main_file: Path) -> None:
            super().__init__()
            self.main_file = main_file
            self.points = [object()] * 2
            self.texts = [object()]
            self.pr5 = _Pr5Like()

    def _fake_create(main_file: Path) -> Any:
        return _ContentWithPr5(main_file)

    def _fake_export(content: Any, xlsx_path: Path) -> None:
        xlsx_path.write_bytes(b"x" * 150)

    import mapsys.parser.content as mcontent

    monkeypatch.setattr(mcontent.Content, "create", _fake_create)
    monkeypatch.setattr("mapsys.xl.export_to_xlsx", _fake_export)

    report_path = tmp_path / "export_report.xlsx"
    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "to-xlsx-dir",
            str(tmp_path),
            "--report",
            str(report_path),
        ],
    )

    assert result.exit_code == 0
    assert report_path.exists()
    assert "Report written" in result.output
    assert (tmp_path / "proj.xlsx").exists()

    from openpyxl import load_workbook

    wb = load_workbook(report_path)
    assert "Conversion report" in wb.sheetnames
    ws = wb["Conversion report"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "Mapsys path" in headers
    assert "XLSX path" in headers
    assert "points" in headers
    assert ws.max_row >= 2
    assert ws.cell(row=2, column=3).value == "proj"
    assert ws.cell(row=2, column=4).value == 2
    assert ws.cell(row=2, column=14).value == 150


def test_cli_to_xlsx_dir_report_empty_when_no_exports(
    tmp_path: Path, monkeypatch: Any
) -> None:
    """`to-xlsx-dir --report` creates report with headers only when no .pr5."""
    report_path = tmp_path / "report.xlsx"
    runner = CliRunner()
    result = runner.invoke(
        cli,
        [
            "to-xlsx-dir",
            str(tmp_path),
            "--report",
            str(report_path),
        ],
    )
    assert result.exit_code == 0
    assert report_path.exists()
    from openpyxl import load_workbook

    wb = load_workbook(report_path)
    ws = wb["Conversion report"]
    assert ws.max_row == 1
