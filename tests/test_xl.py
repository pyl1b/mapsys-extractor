"""Tests for the XLSX exporter in ``mapsys.xl``.

These tests construct a minimal in-memory ``Content`` instance and verify
that the generated workbook contains the expected sheets, columns and basic
formatting (notably the 3-decimal format for floats).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from mapsys.parser.al5_poly_layer import Al5Data
from mapsys.parser.ar5_polys import Ar5Data
from mapsys.parser.content import Content
from mapsys.parser.n05_points import No5Coord
from mapsys.parser.te5_text_meta import Te5TextMeta
from mapsys.parser.ts5_text_store import Ts5Text
from mapsys.xl import export_to_xlsx, write_dxf_report, write_xlsx_report


def _make_min_content(tmp_path: Path) -> Content:
    """Create a minimal Content with a few rows for each table.

    The PR5 portion is intentionally omitted to exercise the branch where
    PR5 is not present.
    """

    main = tmp_path / "proj.pr5"
    main.write_bytes(b"dummy")

    content = Content(main_file=main, files={})

    # NO5 points: one record with distinct float fields
    content.points = [
        No5Coord(
            type=16,
            id_nr=1,
            layer=2,
            pt_nr=100,
            east=123.4567,
            north=765.4321,
            z=10.5,
            uniq=42,
            connexion=0,
        )
    ]

    # TS5 texts and TE5 meta
    content.texts = [Ts5Text(offset=0, text="Hello")]
    content.t_meta = [
        Te5TextMeta(
            first_zero=0,
            text_id=1,
            layer=0,
            font=0,
            flags=0,
            height=1.0,
            direction=0.0,
            east=1.0,
            north=2.0,
            align_east=0.0,
            align_north=0.0,
            z=0.0,
            offset=0,
            length=2,
        )
    ]

    # AR5/AS5/AL5
    content.p_meta = [
        Ar5Data(
            unk8=0,
            line_id=1,
            line_nr=1,
            unk1=0,
            vertex_offset=0,
            vertex_count=2,
            lay_rec=0,
            layer_count=1,
            unk6=0,
            unk7=0,
        )
    ]
    content.v_offsets = [0, 0]
    content.p_layers = [Al5Data(layer=0, second=0, third=0)]

    return content


def test_export_to_xlsx_creates_expected_sheets_and_columns(
    tmp_path: Path,
) -> None:
    """Export creates expected sheets, columns and number formats."""

    content = _make_min_content(tmp_path)
    out = tmp_path / "out.xlsx"
    export_to_xlsx(content, out)

    assert out.exists(), "Workbook file should be created"

    wb = load_workbook(out)
    # Core sheets
    assert "NO5_points" in wb.sheetnames
    assert "TS5_texts" in wb.sheetnames
    assert "TE5_meta" in wb.sheetnames
    assert "AR5_polys" in wb.sheetnames
    assert "AS5_offsets" in wb.sheetnames
    assert "AL5_layers" in wb.sheetnames
    assert "Headers" in wb.sheetnames

    # NO5_points: header row contains idx and east/north/z columns
    ws = wb["NO5_points"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for col in ["idx", "east", "north", "z"]:
        assert col in headers

    # Floats should have 0.000 number format where applicable (east)
    east_col = headers.index("east") + 1
    cell = ws.cell(row=2, column=east_col)
    assert cell.number_format == "0.000"

    # Headers sheet should include NO5 count row
    hs = wb["Headers"]
    rows = list(hs.iter_rows(min_row=1, max_row=6, max_col=2))
    flat = [cell.value for row in rows for cell in row]
    assert "NO5" in flat
    assert "count" in flat


class TestWriteDxfReport:
    """Tests for write_dxf_report."""

    def test_write_dxf_report_creates_workbook_with_headers(
        self, tmp_path: Path
    ) -> None:
        """Report with no rows creates file with headers only."""
        out = tmp_path / "report.xlsx"
        write_dxf_report([], out)
        assert out.exists()
        wb = load_workbook(out)
        assert "Conversion report" in wb.sheetnames
        ws = wb["Conversion report"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "Mapsys path" in headers
        assert "DXF path" in headers
        assert "Mapsys name" in headers
        assert "points" in headers
        assert "Generated at" in headers
        assert "DXF size (bytes)" in headers

    def test_write_dxf_report_includes_hyperlinks(
        self, tmp_path: Path
    ) -> None:
        """Report row includes hyperlinks for file paths."""
        import datetime

        mapsys_path = (tmp_path / "proj.pr5").as_posix()
        dxf_path = (tmp_path / "proj.dxf").as_posix()
        rows = [
            {
                "mapsys_path": mapsys_path,
                "dxf_path": dxf_path,
                "mapsys_name": "proj",
                "points": 5,
                "texts": 2,
                "t_meta": 2,
                "p_meta": 1,
                "v_offsets": 4,
                "p_layers": 1,
                "pr5_layers": 3,
                "pr5_after": 0,
                "pr5_fonts": 1,
                "generated_at": datetime.datetime(2025, 3, 1, 12, 30, 0),
                "dxf_size_bytes": 12345,
            }
        ]
        out = tmp_path / "report.xlsx"
        write_dxf_report(rows, out)
        wb = load_workbook(out)
        ws = wb["Conversion report"]
        mapsys_cell = ws.cell(row=2, column=1)
        assert mapsys_cell.hyperlink is not None
        target = mapsys_cell.hyperlink.target
        assert target is not None and "file:" in target
        dxf_cell = ws.cell(row=2, column=2)
        assert dxf_cell.hyperlink is not None
        dxf_target = dxf_cell.hyperlink.target
        assert dxf_target is not None and "file:" in dxf_target

    def test_write_dxf_report_content_counts(self, tmp_path: Path) -> None:
        """Report row contains correct Content counts and metadata."""
        import datetime

        rows = [
            {
                "mapsys_path": str(tmp_path / "a.pr5"),
                "dxf_path": str(tmp_path / "a.dxf"),
                "mapsys_name": "a",
                "points": 10,
                "texts": 3,
                "t_meta": 3,
                "p_meta": 2,
                "v_offsets": 6,
                "p_layers": 2,
                "pr5_layers": 0,
                "pr5_after": 0,
                "pr5_fonts": 0,
                "generated_at": datetime.datetime(2025, 1, 15, 9, 0, 0),
                "dxf_size_bytes": 999,
            }
        ]
        out = tmp_path / "report.xlsx"
        write_dxf_report(rows, out)
        wb = load_workbook(out)
        ws = wb["Conversion report"]
        assert ws.cell(row=2, column=3).value == "a"
        assert ws.cell(row=2, column=4).value == 10
        assert ws.cell(row=2, column=5).value == 3
        assert ws.cell(row=2, column=14).value == 999
        assert "2025-01-15 09:00:00" in str(ws.cell(row=2, column=13).value)


class TestWriteXlsxReport:
    """Tests for write_xlsx_report."""

    def test_write_xlsx_report_creates_workbook_with_headers(
        self, tmp_path: Path
    ) -> None:
        """Report with no rows creates file with headers including XLSX path."""
        out = tmp_path / "report.xlsx"
        write_xlsx_report([], out)
        assert out.exists()
        wb = load_workbook(out)
        assert "Conversion report" in wb.sheetnames
        ws = wb["Conversion report"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "XLSX path" in headers
        assert "XLSX size (bytes)" in headers
