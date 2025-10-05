"""XLSX export utilities.

This module exports parsed ``mapsys`` content into an Excel workbook.

Design goals:

- One worksheet per logical table in :class:`mapsys.parser.content.Content`.
- Each sheet contains a single Excel Table covering the header row and all
  data rows.
- Nested structures (dataclasses, tuples) are flattened into individual
  columns using stable, descriptive column names.
- An extra ``idx`` column is added to each table with the 0-based row index.
- A separate ``Headers`` worksheet lists file headers/top-level metadata in
  sections, in source order, with one empty row between sections; it does not
  contain an Excel Table.
- Cell types are set appropriately where known. Unknown types are written as
  text. Floating-point numbers are formatted with 3 decimal places.

"""

from __future__ import annotations

import dataclasses
import logging
from collections.abc import Iterable
from dataclasses import is_dataclass
from pathlib import Path
from typing import TYPE_CHECKING, Any, cast

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

if TYPE_CHECKING:  # import only for type checking to avoid runtime deps
    from mapsys.parser.content import Content as _Content

logger = logging.getLogger(__name__)


#
# Utilities
#


def _is_primitive(value: Any) -> bool:
    """Return True for simple scalar types we can write directly.

    Known primitives include: None, bool, int, float, str, bytes.
    """

    return value is None or isinstance(value, (bool, int, float, str, bytes))


def _flatten_value(prefix: str, value: Any, out: dict[str, Any]) -> None:
    """Flatten ``value`` into ``out`` using ``prefix`` for column names.

    - Dataclasses: expand each field with ``prefix_field``.
    - Tuples/lists: expand indexed as ``prefix_i``.
    - Bytes: store as hex string to make it textual and compact.
    - Primitives: store directly.
    - Unknown objects: store their string representation.
    """

    # Guard: primitives
    if _is_primitive(value):
        if isinstance(value, bytes):
            out[prefix] = value.hex()
        else:
            out[prefix] = value
        return

    # Dataclass
    if is_dataclass(value):
        for field in dataclasses.fields(value):
            _flatten_value(
                f"{prefix}_{field.name}", getattr(value, field.name), out
            )
        return

    # Mapping-like (skip to avoid spreading arbitrary dicts; stringify)
    if isinstance(value, dict):
        out[prefix] = str(value)
        return

    # Iterable (list/tuple) of scalars/records → index columns
    if isinstance(value, Iterable) and not isinstance(value, (str, bytes)):
        idx = 0
        for item in value:
            _flatten_value(f"{prefix}_{idx}", item, out)
            idx += 1
        # If empty iterable, record an empty marker for visibility
        if idx == 0 and prefix not in out:
            out[prefix] = None
        return

    # Fallback: stringify
    out[prefix] = str(value)


def _row_dict_from_obj(index: int, obj: Any) -> dict[str, Any]:
    """Create a flattened row dict for ``obj`` including ``idx`` column."""

    row: dict[str, Any] = {"idx": index}

    # Primitive at top-level: store under a generic 'value' column.
    if _is_primitive(obj):
        if isinstance(obj, bytes):
            row["value"] = obj.hex()
        else:
            row["value"] = obj
        return row

    # Decide object kind
    if is_dataclass(obj):
        for field in dataclasses.fields(obj):
            value = getattr(obj, field.name)
            _flatten_value(field.name, value, row)
    elif isinstance(obj, dict):
        for key, value in obj.items():
            _flatten_value(str(key), value, row)
    else:
        # Generic object: reflect attributes conservatively
        for key in dir(obj):
            if key.startswith("_"):
                continue
            try:
                value = getattr(obj, key)
            except Exception:
                continue
            if callable(value):
                continue
            _flatten_value(key, value, row)

    return row


def _write_table_sheet(
    wb: Workbook,
    title: str,
    objects: list[Any],
) -> None:
    """Create a worksheet named ``title`` with a single Excel table.

    The function infers columns from the union of keys across all rows to
    guarantee a stable schema and writes values row by row. Numeric types are
    written as numbers; strings as text. Floats receive a 3-decimal format.
    """

    ws = wb.create_sheet(title)

    # Build flattened rows
    rows = [_row_dict_from_obj(i, obj) for i, obj in enumerate(objects)]

    # Column ordering: stable sorted keys
    columns: list[str] = sorted({k for row in rows for k in row.keys()})
    if not columns:
        columns = ["idx"]

    # Write header
    ws.append(columns)

    # Write data
    for row in rows:
        ws.append([row.get(col, None) for col in columns])

    # Apply number formats for floats
    for col_idx, col_name in enumerate(columns, start=1):
        # Heuristic: if any value in column is a float, set number format
        is_float = any(isinstance(r.get(col_name), float) for r in rows)
        if is_float:
            for cell in ws.iter_cols(
                min_col=col_idx, max_col=col_idx, min_row=2
            ):
                for c in cell:
                    c.number_format = "0.000"

    # Create Excel Table covering the data region if any rows
    last_row = ws.max_row
    last_col = ws.max_column
    if last_row >= 1 and last_col >= 1:
        ref = f"A1:{get_column_letter(last_col)}{last_row}"
        table = Table(displayName=f"Tbl_{title}", ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)


def _append_headers_section(
    ws: Worksheet,
    title: str,
    rows: list[tuple[str, Any]],
) -> None:
    """Append a titled section to the headers sheet.

    Writes a title row, then key/value pairs, then an empty row.
    Values of bytes are written as hex strings. Floats use 3 decimals.
    """

    start_row = (ws.max_row + 1) if ws.max_row else 1
    ws.cell(row=start_row, column=1, value=title)
    for key, value in rows:
        start_row += 1
        v = value
        if isinstance(v, bytes):
            v = v.hex()
        ws.cell(row=start_row, column=1, value=key)
        cell = ws.cell(
            row=start_row,
            column=2,
            value=v,
        )
        if isinstance(v, float):
            cell.number_format = "0.000"
    # Empty line separator
    ws.cell(row=start_row + 1, column=1, value=None)


def _headers_rows_from_content(
    content: Any,
) -> list[tuple[str, list[tuple[str, Any]]]]:
    """Collect header/metadata rows grouped by source for the Headers sheet.

    We only have PR5 header persisted in ``Content``; other sources expose
    counts.
    """

    groups: list[tuple[str, list[tuple[str, Any]]]] = []

    # NO5/AR5/AS5/AL5/TE5/TS5 headers are not kept on Content; expose counts.

    groups.append(
        (
            "NO5",
            [
                ("count", len(content.points)),
            ],
        )
    )

    groups.append(
        (
            "TS5",
            [
                ("count", len(content.texts)),
            ],
        )
    )

    groups.append(
        (
            "TE5",
            [
                ("count", len(content.t_meta)),
            ],
        )
    )

    groups.append(
        (
            "AR5",
            [
                ("count", len(content.p_meta)),
            ],
        )
    )

    groups.append(
        (
            "AS5",
            [
                ("count", len(content.v_offsets)),
            ],
        )
    )

    groups.append(
        (
            "AL5",
            [
                ("count", len(content.p_layers)),
            ],
        )
    )

    # PR5 header if available
    if getattr(content, "pr5", None) is not None:
        pr5 = content.pr5
        hdr = pr5.head
        rows: list[tuple[str, Any]] = [
            ("file_path", hdr.file_path),
            ("dir_path", hdr.dir_path),
            ("east_min", hdr.east_min),
            ("east_max", hdr.east_max),
            ("north_min", hdr.north_min),
            ("north_max", hdr.north_max),
            ("a_zero", hdr.a_zero),
            ("a_one", hdr.a_one),
        ]
        groups.append(("PR5.Header", rows))

        groups.append(("PR5.Fonts", [("count", len(pr5.font_names))]))

    return groups


def _auto_size_columns(ws: Worksheet) -> None:
    """Auto-size columns based on content width (simple heuristic)."""

    # Use column indices to avoid issues with merged cells in header row.
    max_col = int(ws.max_column or 0)
    max_row = int(ws.max_row or 0)
    for col_idx in range(1, max_col + 1):
        max_length = 0
        for cells in ws.iter_cols(
            min_col=col_idx, max_col=col_idx, min_row=1, max_row=max_row
        ):
            for cell in cells:
                value = cell.value
                length = len(str(value)) if value is not None else 0
                if length > max_length:
                    max_length = length
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(
            60, max(8, max_length + 2)
        )


#
# Public API
#


def export_to_xlsx(content: Any, xlsx_path: Path) -> None:
    """Export ``content`` into an XLSX file at ``xlsx_path``.

    The function creates or overwrites the file.
    """

    if TYPE_CHECKING:
        assert isinstance(content, _Content)

    wb = Workbook()
    # Remove default sheet to keep only named ones
    default = wb.active
    default_ws: Worksheet = cast(Worksheet, default)
    wb.remove(default_ws)

    # Data sheets: one per table
    _write_table_sheet(wb, "NO5_points", list(content.points))
    _write_table_sheet(wb, "TS5_texts", list(content.texts))
    _write_table_sheet(wb, "TE5_meta", list(content.t_meta))
    _write_table_sheet(wb, "AR5_polys", list(content.p_meta))
    _write_table_sheet(wb, "AS5_offsets", list(content.v_offsets))
    _write_table_sheet(wb, "AL5_layers", list(content.p_layers))

    # PR5 supplemental tables
    if content.pr5 is not None:
        pr5 = content.pr5
        _write_table_sheet(wb, "PR5_layers", list(pr5.layers))
        _write_table_sheet(wb, "PR5_after", list(pr5.after))
        _write_table_sheet(wb, "PR5_fonts", list(pr5.font_names))

    # Headers sheet
    ws_headers = wb.create_sheet("Headers")
    for title, rows in _headers_rows_from_content(content):
        _append_headers_section(ws_headers, title, rows)

    # Auto-size all sheets
    for sheet in wb.worksheets:
        _auto_size_columns(sheet)

    # Save file
    wb.save(xlsx_path.as_posix())


__all__ = ["export_to_xlsx"]
